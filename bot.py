import telebot
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
from flask import Flask
import threading
import os
import time
import requests
import openpyxl
from io import BytesIO

# ==========================================
# 1. BOT SETTINGS
# ==========================================
BOT_TOKEN = "8665786518:AAHzrG19WCAu-AuTt1LZBeqm446eoV0n-zs"
bot = telebot.TeleBot(BOT_TOKEN)

SECRET_CODE = "dkstudio"
unlocked_users = set()

# Aapki Google Sheet ka URL (Excel format me download karne ke liye)
SHEET_URL = "https://docs.google.com/spreadsheets/d/1eMdE5uhdQXpA73_NO6MwJbMq7x_eI7C38jbqjzP1vtY/export?format=xlsx"

# Bot ka dimag jisme sheet ka data aayega
courses_data = {}
user_quiz_state = {}

# ==========================================
# 2. GOOGLE SHEET AUTO-FETCHER
# ==========================================
def update_courses_from_sheet():
    global courses_data
    try:
        print("Google Sheet se naya data laa raha hu...")
        response = requests.get(SHEET_URL)
        workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
        
        new_data = {}
        chap_num = 1
        
        # Har naye Tab (Sheet) ko padhna
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            questions = []
            
            # Row 2 se padhna shuru karega (Row 1 me Heading honi chahiye)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None and str(row[0]).strip() != "":
                    try:
                        # Column 0: Question, 1-4: Options, 5: Ans, 6: Expl
                        q_text = str(row[0])
                        options = [str(row[1]), str(row[2]), str(row[3]), str(row[4])]
                        
                        # Answer ko 1,2,3,4 ya A,B,C,D se set karna
                        ans_val = str(row[5]).strip().upper()
                        if ans_val in ['A', '1']: ans_idx = 0
                        elif ans_val in ['B', '2']: ans_idx = 1
                        elif ans_val in ['C', '3']: ans_idx = 2
                        elif ans_val in ['D', '4']: ans_idx = 3
                        else: ans_idx = 0 # Default option A
                        
                        expl = str(row[6]) if row[6] is not None else ""
                        expl += "\n\nJoin: @GKGSCOMPLETEPYQREVISION_bot"
                        
                        questions.append({
                            "q": q_text,
                            "options": options,
                            "ans": ans_idx,
                            "expl": expl
                        })
                    except Exception as e:
                        print(f"Row error skipped: {e}")
            
            if questions:
                chap_key = f"chap_{chap_num}"
                # Pehla tab FREE hoga, baki sab PAID
                is_free = True if chap_num == 1 else False
                icon = "🟢 FREE" if is_free else "🔴 PAID"
                
                new_data[chap_key] = {
                    "title": f"Chapter {chap_num}: {sheet_name} ({icon})",
                    "is_free": is_free,
                    "questions": questions
                }
                chap_num += 1
                
        courses_data = new_data
        return True, f"✅ Successfully Updated!\nTotal Chapters: {len(courses_data)}"
    except Exception as e:
        return False, f"❌ Error: {e}"

# Server start hote hi sheet ko fetch kar lo
update_courses_from_sheet()

# ==========================================
# 3. DUMMY WEB SERVER (For Render)
# ==========================================
app = Flask(__name__)

@app.route('/')
def home():
    return f"Auto-Sync Quiz Bot Live! Chapters Loaded: {len(courses_data)}"

# ==========================================
# 4. BOT LOGIC (OFFICIAL QUIZ MODE)
# ==========================================
@bot.message_handler(commands=['start'])
def send_welcome(message):
    chat_id = message.chat.id
    if chat_id in user_quiz_state:
        del user_quiz_state[chat_id]

    welcome_text = (
        "🤖 *Welcome to Official BPSC Quiz Bot!*\n\n"
        "🟢 *Chapter 1* bilkul FREE hai.\n"
        "🔒 *Paid Chapters* unlock karne ke liye apna Access Code bhejein (Example: `dkstudio`).\n\n"
        "*(Admins can type /update to refresh questions)*"
    )
    
    markup = InlineKeyboardMarkup(row_width=1)
    if not courses_data:
        bot.send_message(chat_id, "⏳ Sheet se questions load ho rahe hain... Thodi der me /start dabayein.")
        return
        
    for chap_id, chap_info in courses_data.items():
        icon = "🟢" if chap_info['is_free'] else ("📖" if chat_id in unlocked_users else "🔒")
        btn = InlineKeyboardButton(f"{icon} {chap_info['title']}", callback_data=f"menu_{chap_id}")
        markup.add(btn)
        
    bot.send_message(chat_id, welcome_text, reply_markup=markup, parse_mode="Markdown")

@bot.message_handler(commands=['update'])
def force_update(message):
    bot.send_message(message.chat.id, "⏳ Google Sheet se naye questions fetch kar raha hu...")
    success, msg = update_courses_from_sheet()
    bot.send_message(message.chat.id, msg)

@bot.message_handler(func=lambda message: True)
def handle_text(message):
    chat_id = message.chat.id
    text = message.text.strip().lower()
    
    if text == SECRET_CODE.lower():
        unlocked_users.add(chat_id)
        bot.send_message(chat_id, "✅ *Success!* Pura course unlock ho gaya hai.\nAb /start dabayein.", parse_mode="Markdown")
    else:
        bot.send_message(chat_id, "❌ *Invalid Code!*")

@bot.callback_query_handler(func=lambda call: call.data.startswith("menu_"))
def handle_menu(call):
    chat_id = call.message.chat.id
    chap_id = call.data.split("_")[1]
    
    if chap_id not in courses_data:
        return
        
    if not courses_data[chap_id]["is_free"] and chat_id not in unlocked_users:
        bot.send_message(chat_id, "🔒 *Chapter Locked!*\nIse access karne ke liye Access Code bhejein.", parse_mode="Markdown")
        return

    user_quiz_state[chat_id] = {"chap": chap_id, "q_idx": 0}
    bot.send_message(chat_id, f"🚀 *Starting {courses_data[chap_id]['title']}...*", parse_mode="Markdown")
    send_next_poll(chat_id)

def send_next_poll(chat_id):
    state = user_quiz_state.get(chat_id)
    if not state: return
    
    chap_id = state["chap"]
    q_idx = state["q_idx"]
    questions = courses_data[chap_id]["questions"]
    
    if q_idx < len(questions):
        q = questions[q_idx]
        bot.send_poll(
            chat_id=chat_id,
            question=f"Q{q_idx+1}: {q['q']}",
            options=q["options"],
            type='quiz',
            correct_option_id=q["ans"],
            explanation=q["expl"],
            is_anonymous=False
        )
    else:
        is_free = courses_data[chap_id]["is_free"]
        msg = "🏁 *Quiz Khatam!*"
        if is_free and chat_id not in unlocked_users:
            msg += "\n\nAgar demo pasand aaya toh pura course unlock karne ke liye code `dkstudio` bhejein."
        bot.send_message(chat_id, msg, parse_mode="Markdown")
        if chat_id in user_quiz_state: del user_quiz_state[chat_id]

@bot.poll_answer_handler(func=lambda answer: True)
def handle_poll_answer(answer):
    chat_id = answer.user.id
    state = user_quiz_state.get(chat_id)
    if state:
        state["q_idx"] += 1
        time.sleep(1.5)
        send_next_poll(chat_id)

# ==========================================
# 5. STARTING ENGINE
# ==========================================
def run_bot():
    while True:
        try: bot.polling(none_stop=True, interval=0, timeout=20)
        except Exception as e: time.sleep(5)

if __name__ == "__main__":
    bot_thread = threading.Thread(target=run_bot)
    bot_thread.start()
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
